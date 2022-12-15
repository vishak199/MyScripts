import os
from glob import glob
import pandas as pd
import os.path
import sys
from datetime import datetime
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook

PATH = sys.argv[1]
PATHDATE = sys.argv[2]
Person_Path = PATH +r'/Person.xlsx'
Person1_Path = PATH + r'/Person1.xlsx'
PersonGroup_Path = PATH + r'/PersonGroup.xlsx'
PersonGroup1_Path = PATH + r'/PersonGroup1.xlsx'
PersonToGroup_Path = PATH + r'/PersonToGroup.xlsx'
company_Path = PATH + r'/Company.xlsx'
Location_Path = PATH + r'/Location.xlsx'
FinalPerson_Path = PATH + r'/Person.csv'
FinalPersonGroup_Path = PATH + r'/Group.csv'
FinalPersonToGroup_Path = PATH + r'/Person_To_Group.csv'
now = datetime.now()
x = now.strftime("%Y_%m_%d_%H-%M-%S")
FinalPersonDate_Path= PATHDATE + r'/Person'+ "_" + x + ".csv"
FinalPersonGroupDate_Path= PATHDATE + r'/Group'+ "_" + x + ".csv"
FinalPersonToGroupDate_Path= PATHDATE + r'/Person_To_Group'+ "_" + x + ".csv"
EXT = "*.zip"
all_zip_files = [file
                 for path, subdir, files in os.walk(PATH)
                 for file in glob(os.path.join(path, EXT))]
for x in all_zip_files:
    os.remove(x)
    print("zip deleted")
################################################Person_ File ################################################
EXT1 = "Person_*.csv"
all_person_files = [file
                 for path, subdir, files in os.walk(PATH)
                 for file in glob(os.path.join(path, EXT1))]
df_from_person_file = (pd.read_csv(f, sep=',') for f in all_person_files)
df_person_merged   = pd.concat(df_from_person_file, ignore_index=True)
df_person_merged.drop(df_person_merged.columns[[0, 1]], axis = 1, inplace = True)
df=df_person_merged.to_excel(Person_Path, index=False)
df=df_person_merged.to_excel(Person1_Path, index=False)
wb=load_workbook(Person_Path)
sheet=wb.active
for y in all_person_files:
    os.remove(y)
    print("person merge done")
##############################################PersonGroup_ File ##############################################
EXT2 = "PersonGroup_*.csv"
all_persongroup_files = [file
                 for path, subdir, files in os.walk(PATH)
                 for file in glob(os.path.join(path, EXT2))]
df_from_persongroup_file = (pd.read_csv(f, sep=',') for f in all_persongroup_files)
df_persongroup_merged   = pd.concat(df_from_persongroup_file, ignore_index=True)
df_persongroup_merged.drop(df_persongroup_merged.columns[[0, 1]], axis = 1, inplace = True)
df=df_persongroup_merged.to_excel(PersonGroup_Path, index=False)
wb=load_workbook(PersonGroup_Path)
sheet=wb.active
#persongroup_max_row=sheet.max_row
#print(persongroup_max_row+1)
for z in all_persongroup_files:
    os.remove(z)
##############################################PersonToGroup_ File ##############################################
EXT3 = "PersonToGroup_*.csv"
all_persontogroup_files = [file
                 for path, subdir, files in os.walk(PATH)
                 for file in glob(os.path.join(path, EXT3))]
df_from_persontogroup_file = (pd.read_csv(f, sep=',') for f in all_persontogroup_files)
df_persontogroup_merged   = pd.concat(df_from_persontogroup_file, ignore_index=True)
df_persontogroup_merged.drop(df_persontogroup_merged.columns[[0, 1]], axis = 1, inplace = True)
df=df_persontogroup_merged.to_excel(PersonToGroup_Path, index=False)
wb = load_workbook(PersonToGroup_Path)
ws = wb.active
#ws['C1'] = 'Group_Name'
#ws['D1'] = 'Person_Name'
wb.save(PersonToGroup_Path)
for p in all_persontogroup_files:
    os.remove(p)
##########################################Company File #############################################################
EXT4 = "Company_*.csv"
all_company_files = [file
                 for path, subdir, files in os.walk(PATH)
                 for file in glob(os.path.join(path, EXT4))]
df_from_company_file = (pd.read_csv(f, sep=',') for f in all_company_files)
df_company_merged   = pd.concat(df_from_company_file, ignore_index=True)
df_company_merged.drop(df_company_merged.columns[[0, 1]], axis = 1, inplace = True)
dict = {'Id': 'Company',}
df_company_merged.rename(columns=dict,inplace=True)
df1_company_merged=df_company_merged.append({'Company':'', 'DisplayLabel':'#N/A '}, ignore_index=True)
df=df1_company_merged.to_excel(company_Path, index=False)
wb = load_workbook(company_Path)
ws = wb.active
wb.save(company_Path)
for p in all_company_files:
    os.remove(p)

###################################Location File ########################################################################
EXT6 = "Location_*.csv"
all_Location_files = [file
                 for path, subdir, files in os.walk(PATH)
                 for file in glob(os.path.join(path, EXT6))]
df_from_Location_file = (pd.read_csv(f, sep=',') for f in all_Location_files)
df_Location_merged   = pd.concat(df_from_Location_file, ignore_index=True)
df_Location_merged.drop(df_Location_merged.columns[[0, 1, 2, 3]], axis = 1, inplace = True)
dict = {'Id': 'Location','Name': 'Location Name'}
df_Location_merged.rename(columns=dict,inplace=True)
df1_Location_merged=df_Location_merged.append({'Location':'', 'Location Name':'#N/A '}, ignore_index=True)
df=df1_Location_merged.to_excel(Location_Path, index=False)
wb = load_workbook(Location_Path)
ws = wb.active
wb.save(Location_Path)
for p in all_Location_files:
    os.remove(p)

#################################################Person File VLOOKUP#####################################################
df1 = pd.read_excel(Person_Path,engine='openpyxl')
df2 = pd.read_excel(company_Path , engine='openpyxl')
             ############For Company##################################
join = pd.merge(df1,df2[['Company', 'DisplayLabel']],on = 'Company' , how = 'left')
join.drop(join.columns[[0]], axis = 1, inplace = True)
dict = {'DisplayLabel': 'Company',}
join.rename(columns=dict,inplace=True)
column = join.pop('Company')
join.insert(0, 'Company' , column)
df3 = join.to_excel(Person_Path,engine='openpyxl',index=False)
                         ################For Location####################
df1 = pd.read_excel(Person_Path,engine='openpyxl')
df2 = pd.read_excel(Location_Path , engine='openpyxl')
join = pd.merge(df1,df2[['Location', 'Location Name']],on = 'Location' , how = 'left')
join.drop(join.columns[[13]], axis = 1, inplace = True)
dict = {'Location Name': 'Location',}
join.rename(columns=dict,inplace=True)
column = join.pop('Location')
join.insert(13, 'Location' , column)
df3 = join.to_excel(Person_Path,engine='openpyxl',index=False)
                     ####################For Manager##################
df1 = pd.read_excel(Person1_Path,engine='openpyxl')
df1.drop(df1.columns[[0,1,3,4,5,6,7,8,9,10,11,12,13,14,15,16,18,19,20,21]], axis = 1, inplace = True)
dict = {'Id': 'Manager','Name': 'Manager Name'}
df1.rename(columns=dict,inplace=True)
df2=df1.append({'Manager':'', 'Manager Name':'#N/A '}, ignore_index=True)
df = df2.to_excel(Person1_Path,engine='openpyxl',index=False)
df1 = pd.read_excel(Person_Path,engine='openpyxl')
df2 = pd.read_excel(Person1_Path, engine='openpyxl')
join = pd.merge(df1,df2[['Manager', 'Manager Name']],on = 'Manager' , how = 'left')
join.drop(join.columns[[14]], axis = 1, inplace = True)
dict = {'Manager Name': 'Manager',}
join.rename(columns=dict,inplace=True)
column = join.pop('Manager')
join.insert(14, 'Manager' , column)
df3 = join.to_excel(Person_Path,engine='openpyxl',index=False)


##########################################################Person To Group #############################
df1=pd.read_excel(PersonGroup_Path,engine='openpyxl')
df1.drop(df1.columns[[2,3]], axis = 1, inplace = True)
dict = {'Id': 'secondEntity_PersonGroup','Name': 'Group Name'}
df1.rename(columns=dict,inplace=True)
df2=df1.append({'secondEntity_PersonGroup':'', 'Group Name':'#N/A '}, ignore_index=True)
df = df2.to_excel(PersonGroup1_Path,engine='openpyxl',index=False)
df3 = pd.read_excel(Person1_Path,engine='openpyxl')
dict = {'Manager':'firstEntity_Person','Manager Name':'Person Name'}
df3.rename(columns=dict,inplace=True)
#df4=df3.append({'firstEntity_Person':'', 'Person Name':'#N/A '}, ignore_index=True)
df = df3.to_excel(Person1_Path,engine='openpyxl',index=False)
                #############################Person Name Vlookup#####################
df1 = pd.read_excel(PersonToGroup_Path,engine='openpyxl')
df2 = pd.read_excel(Person1_Path, engine='openpyxl')
join = pd.merge(df1,df2[['firstEntity_Person', 'Person Name']],on = 'firstEntity_Person' , how = 'left')
column = join.pop('Person Name')
join.insert(1, 'Person Name' , column)
df3 = join.to_excel(PersonToGroup_Path,engine='openpyxl',index=False)
             ########################Group Name Vlookup#####################
df1 = pd.read_excel(PersonToGroup_Path,engine='openpyxl')
df2 = pd.read_excel(PersonGroup1_Path, engine='openpyxl')
join = pd.merge(df1,df2[['secondEntity_PersonGroup', 'Group Name']],on = 'secondEntity_PersonGroup' , how = 'left')
column = join.pop('Group Name')
join.insert(3, 'Group Name' , column)
df3 = join.to_excel(PersonToGroup_Path,engine='openpyxl',index=False)
os.remove(PersonGroup1_Path)
             #########################Group Owner Vlookup#####################
df1=pd.read_excel(PersonGroup_Path,engine='openpyxl')
df1.drop(df1.columns[[1,2]], axis = 1, inplace = True)
dict = {'Id': 'secondEntity_PersonGroup','Owner': 'Group Owner'}
df1.rename(columns=dict,inplace=True)
df2=df1.append({'secondEntity_PersonGroup':'', 'Group Owner':'#N/A '}, ignore_index=True)
df = df2.to_excel(PersonGroup1_Path,engine='openpyxl',index=False)
df1 = pd.read_excel(PersonToGroup_Path,engine='openpyxl')
df2 = pd.read_excel(PersonGroup1_Path, engine='openpyxl')
join = pd.merge(df1,df2[['secondEntity_PersonGroup', 'Group Owner']],on = 'secondEntity_PersonGroup' , how = 'left')
column = join.pop('Group Owner')
join.insert(4, 'Group Owner' , column)
df3 = join.to_excel(PersonToGroup_Path,engine='openpyxl',index=False)
os.remove(PersonGroup1_Path)
             ##########################Group Owner Name Vlookup#####################
df1 = pd.read_excel(Person1_Path,engine='openpyxl')
dict = {'firstEntity_Person': 'Group Owner','Person Name': 'Group Owner Name'}
df1.rename(columns=dict,inplace=True)
#df2=df1.append({'Group Owner':'', 'Group Owner Name':'#N/A '}, ignore_index=True)
df = df1.to_excel(Person1_Path,engine='openpyxl',index=False)
df1 = pd.read_excel(PersonToGroup_Path,engine='openpyxl')
df2 = pd.read_excel(Person1_Path, engine='openpyxl')
join = pd.merge(df1,df2[['Group Owner', 'Group Owner Name']],on = 'Group Owner' , how = 'left')
column = join.pop('Group Owner Name')
join.insert(5, 'Group Owner Name' , column)
df3 = join.to_excel(PersonToGroup_Path,engine='openpyxl',index=False)
os.remove(Person1_Path)
              ###########################Group Owner Employee ID Vlookup####################
df1 = pd.read_excel(Person_Path,engine='openpyxl')
df1.drop(df1.columns[[0,1,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21]], axis = 1, inplace = True)
dict = {'Id': 'Group Owner','EmployeeNumber': 'Group Owner Employee ID'}
df1.rename(columns=dict,inplace=True)
df2=df1.append({'Group Owner':'', 'Group Owner Employee ID':'#N/A '}, ignore_index=True)
df = df2.to_excel(Person1_Path,engine='openpyxl',index=False)
df1 = pd.read_excel(PersonToGroup_Path,engine='openpyxl')
df2 = pd.read_excel(Person1_Path, engine='openpyxl')
join = pd.merge(df1,df2[['Group Owner', 'Group Owner Employee ID']],on = 'Group Owner' , how = 'left')
column = join.pop('Group Owner Employee ID')
join.insert(6, 'Group Owner Employee ID' , column)
df3 = join.to_excel(PersonToGroup_Path,engine='openpyxl',index=False)

os.remove(Person1_Path)
               #############################Group Status VLOOKUP###############################
df1 = pd.read_excel(PersonGroup_Path,engine='openpyxl')
df1.drop(df1.columns[[1,3]], axis = 1, inplace = True)
dict = {'Id': 'secondEntity_PersonGroup','Status': 'Group Status'}
df1.rename(columns=dict,inplace=True)
df2=df1.append({'secondEntity_PersonGroup':'', 'Group Status':'#N/A '}, ignore_index=True)
df = df2.to_excel(PersonGroup1_Path,engine='openpyxl',index=False)
df1 = pd.read_excel(PersonToGroup_Path,engine='openpyxl')
df2 = pd.read_excel(PersonGroup1_Path, engine='openpyxl')
join = pd.merge(df1,df2[['secondEntity_PersonGroup', 'Group Status']],on = 'secondEntity_PersonGroup' , how = 'left')
column = join.pop('Group Status')
join.insert(7, 'Group Status' , column)
df3 = join.to_excel(PersonToGroup_Path,engine='openpyxl',index=False)
os.remove(PersonGroup1_Path)
os.remove(company_Path)
os.remove(Location_Path)
######################################################## Group Sheet #######################################################
df1 = pd.read_excel(Person_Path,engine='openpyxl')
df1.drop(df1.columns[[0,1,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21]], axis = 1, inplace = True)
dict = {'Id': 'Owner'}
df1.rename(columns=dict,inplace=True)
df2=df1.append({'Owner':'', 'EmployeeNumber':'#N/A '}, ignore_index=True)
df = df2.to_excel(Person1_Path,engine='openpyxl',index=False)


df1 = pd.read_excel(PersonGroup_Path,engine='openpyxl')
df2 = pd.read_excel(Person1_Path,engine='openpyxl')
join = pd.merge(df1,df2[['Owner', 'EmployeeNumber']],on = 'Owner' , how = 'left')
column = join.pop('EmployeeNumber')
join.insert(3, 'EmployeeNumber' , column)
join=join.loc[(join["Owner"].notnull())]
join=join.loc[(join["Status"]=="Active")]
df = join.to_excel(PersonGroup_Path,engine='openpyxl',index=False)
os.remove(Person1_Path)
#######################################################################Person To Group#########################################
df1 = pd.read_excel(Person_Path,engine='openpyxl')
df1.drop(df1.columns[[0,1,3,4,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21]], axis = 1, inplace = True)
dict = {'Id': 'firstEntity_Person'}
df1.rename(columns=dict,inplace=True)
df2=df1.append({'firstEntity_Person':'', 'EmployeeStatus':'#N/A '}, ignore_index=True)
df = df1.to_excel(Person1_Path,engine='openpyxl',index=False)
df1 = pd.read_excel(PersonToGroup_Path,engine='openpyxl')
df2 = pd.read_excel(Person1_Path,engine='openpyxl')
join = pd.merge(df1,df2,on = 'firstEntity_Person' , how = 'inner')
column = join.pop('EmployeeStatus')
join.insert(2, 'EmployeeStatus' , column)
join=join.loc[(join["EmployeeStatus"]=="Active") | (join["EmployeeStatus"].isnull())]
join=join.loc[(join["Group Status"]=="Active")]
df = join.to_excel(PersonToGroup_Path,index=False)
 ########################### Active values from Person sheet#################
df1 = pd.read_excel(Person_Path,engine='openpyxl',converters={'Id': lambda x: '{0:0>8}'.format(x) ,'EmployeeNumber': lambda x: '{0:0>8}'.format(x)})
df1=df1.loc[(df1["EmployeeStatus"]=="Active") | (df1["EmployeeStatus"].isnull())]
dict = {'EmsCreationTime': 'creationtime'}
df1.rename(columns=dict,inplace=True)
df1=df1[["Company","Email","EmployeeNumber","EmployeeType","EmployeeStatus","creationtime","FirstName","Gender","HireDate","Id","IsDBUser","IsMaasUser","LastName","Location","Manager","MiddleName","MobilePhoneNumber","Name","OfficePhoneNumber","StackIdentifier_c","Title","Upn"]]
epoch_t = df1.creationtime
epoch_hire = df1.HireDate
creationtime=pd.to_datetime(epoch_t, unit='ms').dt.strftime('%m/%d/%Y')
HireDate = pd.to_datetime(epoch_hire, unit='ms').dt.strftime('%m/%d/%Y')
df1['creationtime'] = creationtime
df1['HireDate'] = HireDate
df1['Name'] = df1['Name'].replace('\n','', regex=True)
df1['EmployeeNumber'] = df1['EmployeeNumber'].replace('00000000','', regex=True)
df1.to_csv(FinalPerson_Path, index=False)
df1.to_csv(FinalPersonDate_Path, index=False)
print("final person file created")
################################################# 8 digit conversation of Group csv############################################################
df1 = pd.read_excel(PersonGroup_Path,engine='openpyxl',converters={'Id': lambda x: '{0:0>8}'.format(x) ,'EmployeeNumber': lambda x: '{0:0>8}'.format(x)})
dict = {'EmployeeNumber': 'Owner Employee Id'}
df1.rename(columns=dict,inplace=True)
df1.drop(df1.columns[[2,4]], axis = 1, inplace = True)
df = df1.to_csv(FinalPersonGroup_Path, index=False)
df = df1.to_csv(FinalPersonGroupDate_Path, index=False)
print("final group file created")
################################################# 8 digit conversation of PersonToGroup csv####################################################

df2 = pd.read_excel(PersonToGroup_Path,engine='openpyxl',converters={'firstEntity_Person': lambda x: '{0:0>8}'.format(x) ,'secondEntity_PersonGroup': lambda x: '{0:0>8}'.format(x),'Group Owner': lambda x: '{0:0>8}'.format(x),'Group Owner Employee ID': lambda x: '{0:0>8}'.format(x)})
df2.drop(df2.columns[[2,5,7]], axis = 1, inplace = True)
dict = {'firstEntity_Person': 'Person Id','secondEntity_PersonGroup': 'Group_Id','Group Owner Name': 'Group Owner'}
df2.rename(columns=dict,inplace=True)
df3 = df2.to_csv(FinalPersonToGroup_Path, index=False)
df3 = df2.to_csv(FinalPersonToGroupDate_Path, index=False)
print("final person to group file created")
os.remove(Person_Path)
os.remove(Person1_Path)
os.remove(PersonGroup_Path)
os.remove(PersonToGroup_Path)
